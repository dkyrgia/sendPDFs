import openpyxl
import os

pdfList = []
file_obj_csv = "send_PDFs.csv"
file_obj_xlsx = "list_names.xlsx"

for name in os.listdir("pdfs\\"):
    if name.endswith(".pdf"):
        pdfList.append(name)

print(len(pdfList))
print(pdfList)

wb_obj = openpyxl.load_workbook(file_obj_xlsx)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)
print(cell_obj.value)
max_cols = sheet_obj.max_column
max_rows = sheet_obj.max_row
print(max_rows)
print(max_cols)

for i in range(1, max_rows):
    # print(i)
    name_obj = sheet_obj.cell(row=i + 1, column=2).value
    mail_obj = sheet_obj.cell(row=i + 1, column=3).value
    oldname = "pdfs\\" + str(i) + ".pdf"
    newname = "pdfs\\" + name_obj + ".pdf"

    os.rename(oldname, newname)

pdfList.insert(0, 0)


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')
